import openpyxl

book = openpyxl.load_workbook("PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
#assigning value to a cell
sheet.cell(row=2, column=2).value = "Dhinesh"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A3'].value)

Dict = {}
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)
            #storing the values from excel into a dictionary
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)