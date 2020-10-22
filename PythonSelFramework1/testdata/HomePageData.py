import openpyxl


class HomePageData:

    test_homepage_data = [{"firstname": "Dhinesh", "lastname": "Babu", "gender": "Male"}, {"firstname": "Nina", "lastname": "Williams", "gender":"Female"}]


    @staticmethod
    def get_test_data(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    print(sheet.cell(row=i, column=j).value)
                    # storing the values from excel into a dictionary
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]